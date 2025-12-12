from flask import Blueprint, request, jsonify, send_file
from models import db, User, StudentScore, TeacherStudent
from jwt_utils import token_required
from datetime import datetime, timedelta
from werkzeug.utils import secure_filename
import os
import json
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import io

grades_bp = Blueprint('grades', __name__, url_prefix='/api/grades')

# 配置
UPLOAD_FOLDER = 'uploads/grades'
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}
MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB

# 创建上传文件夹
os.makedirs(UPLOAD_FOLDER, exist_ok=True)


def allowed_file(filename):
    """检查文件是否允许上传"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


# ==================== 成绩录入 ====================

@grades_bp.route('/add', methods=['POST'])
@token_required
def add_score(payload):
    """
    单条成绩录入
    请求体:
    {
        "student_id": 学生ID,
        "subject": "科目",
        "score": 85.5,
        "full_score": 100,
        "exam_date": "2025-12-07",
        "exam_number": 1,
        "notes": "备注"
    }
    """
    try:
        teacher_id = payload.get('user_id')
        data = request.get_json()
        
        # 验证必填字段
        required_fields = ['student_id', 'subject', 'score', 'exam_date']
        for field in required_fields:
            if field not in data or data[field] is None:
                return jsonify({
                    'code': 400,
                    'message': f'缺少必填字段: {field}'
                }), 400
        
        # 验证学生是否存在且是教师的学生
        student = User.query.get(data['student_id'])
        if not student or student.identity != 'student':
            return jsonify({
                'code': 404,
                'message': '学生不存在'
            }), 404
        
        # 验证教师与学生的关系
        teacher_student = TeacherStudent.query.filter_by(
            teacher_id=teacher_id,
            student_id=data['student_id'],
            deleted_at=None
        ).first()
        
        if not teacher_student:
            return jsonify({
                'code': 403,
                'message': '您没有权限为该学生录入成绩'
            }), 403
        
        # 验证成绩范围
        score = float(data['score'])
        full_score = float(data.get('full_score', 100))
        
        if score < 0 or score > full_score:
            return jsonify({
                'code': 400,
                'message': f'成绩需在0-{full_score}之间'
            }), 400
        
        # 创建成绩记录
        student_score = StudentScore(
            student_id=data['student_id'],
            teacher_id=teacher_id,
            subject=data['subject'],
            score=score,
            full_score=full_score,
            exam_date=datetime.strptime(data['exam_date'], '%Y-%m-%d').date(),
            exam_number=data.get('exam_number', 1),
            notes=data.get('notes', '')
        )
        
        db.session.add(student_score)
        db.session.commit()
        
        return jsonify({
            'code': 200,
            'message': '成绩录入成功',
            'data': student_score.to_dict()
        }), 201
        
    except ValueError as e:
        return jsonify({
            'code': 400,
            'message': f'数据格式错误: {str(e)}'
        }), 400
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'code': 500,
            'message': f'成绩录入失败: {str(e)}'
        }), 500


@grades_bp.route('/import', methods=['POST'])
@token_required
def import_scores(payload):
    """
    批量导入成绩（Excel）
    Excel文件格式:
    | 学号 | 姓名 | 科目 | 成绩 | 满分 | 考试日期 | 考试次数 |
    """
    try:
        teacher_id = payload.get('user_id')
        
        if 'file' not in request.files:
            return jsonify({
                'code': 400,
                'message': '没有找到文件'
            }), 400
        
        file = request.files['file']
        
        if file.filename == '':
            return jsonify({
                'code': 400,
                'message': '文件名为空'
            }), 400
        
        if not allowed_file(file.filename):
            return jsonify({
                'code': 400,
                'message': '不支持的文件格式，仅支持 .xlsx 或 .xls'
            }), 400
        
        # 检查文件大小
        file.seek(0, os.SEEK_END)
        file_size = file.tell()
        file.seek(0)
        
        if file_size > MAX_FILE_SIZE:
            return jsonify({
                'code': 400,
                'message': '文件过大，最大支持10MB'
            }), 400
        
        # 读取Excel文件
        wb = load_workbook(file)
        ws = wb.active
        
        scores = []
        errors = []
        
        # 跳过表头，从第2行开始读取
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                if not row[0]:  # 跳过空行
                    continue
                
                student_username = str(row[0]).strip()
                subject = str(row[2]).strip() if row[2] else '未知'
                score = float(row[3]) if row[3] else None
                full_score = float(row[4]) if row[4] else 100
                exam_date_str = str(row[5]).strip() if row[5] else None
                exam_number = int(row[6]) if row[6] else 1
                
                # 查找学生
                student = User.query.filter_by(
                    username=student_username,
                    identity='student',
                    deleted_at=None
                ).first()
                
                if not student:
                    errors.append(f'第{idx}行: 找不到学号为 {student_username} 的学生')
                    continue
                
                # 验证成绩有效性
                if score is None:
                    errors.append(f'第{idx}行: 缺少成绩数据')
                    continue
                
                if score < 0 or score > full_score:
                    errors.append(f'第{idx}行: 成绩 {score} 超出范围 (0-{full_score})')
                    continue
                
                # 解析考试日期
                try:
                    if exam_date_str:
                        exam_date = datetime.strptime(exam_date_str, '%Y-%m-%d').date()
                    else:
                        exam_date = datetime.utcnow().date()
                except:
                    exam_date = datetime.utcnow().date()
                
                # 验证教师与学生的关系
                teacher_student = TeacherStudent.query.filter_by(
                    teacher_id=teacher_id,
                    student_id=student.id,
                    deleted_at=None
                ).first()
                
                if not teacher_student:
                    errors.append(f'第{idx}行: 您没有权限为学生 {student.fullname} 录入成绩')
                    continue
                
                # 创建成绩记录
                student_score = StudentScore(
                    student_id=student.id,
                    teacher_id=teacher_id,
                    subject=subject,
                    score=score,
                    full_score=full_score,
                    exam_date=exam_date,
                    exam_number=exam_number
                )
                
                scores.append(student_score)
                
            except Exception as e:
                errors.append(f'第{idx}行: {str(e)}')
                continue
        
        # 批量保存
        if scores:
            db.session.add_all(scores)
            db.session.commit()
        
        return jsonify({
            'code': 200,
            'message': f'导入成功，共导入 {len(scores)} 条成绩',
            'data': {
                'imported': len(scores),
                'errors': errors,
                'error_count': len(errors)
            }
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'code': 500,
            'message': f'导入失败: {str(e)}'
        }), 500


@grades_bp.route('/template/download', methods=['GET'])
def download_template():
    """
    下载Excel导入模板
    """
    try:
        # 创建新的工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = '成绩导入'
        
        # 设置列宽
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 12
        
        # 创建表头
        headers = ['学号', '姓名', '科目', '成绩', '满分', '考试日期', '考试次数']
        header_fill = PatternFill(start_color='4A90E2', end_color='4A90E2', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 添加示例数据
        sample_data = [
            ['S001', '刘子涛', '数学', 85.5, 100, '2025-12-07', 1],
            ['S002', '杨紫茙', '英语', 92, 100, '2025-12-07', 1],
            ['S003', '王浪', '物理', 78.5, 100, '2025-12-07', 1],
        ]
        
        for row_idx, row_data in enumerate(sample_data, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 添加说明
        ws.append([])
        ws.append(['说明:'])
        ws.append(['1. 学号: 学生的唯一标识，必填'])
        ws.append(['2. 姓名: 学生的姓名，用于验证（可选）'])
        ws.append(['3. 科目: 考试科目，必填'])
        ws.append(['4. 成绩: 学生的考试成绩，必填'])
        ws.append(['5. 满分: 考试总分，默认为100（可选）'])
        ws.append(['6. 考试日期: 考试日期，格式为 YYYY-MM-DD（可选，默认为当前日期）'])
        ws.append(['7. 考试次数: 第几次考试，默认为1（可选）'])
        
        # 保存到字节流
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            as_attachment=True,
            download_name='成绩导入模板.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return jsonify({
            'code': 500,
            'message': f'下载模板失败: {str(e)}'
        }), 500


# ==================== 成绩查询与管理 ====================

@grades_bp.route('/list', methods=['GET'])
@token_required
def get_scores(payload):
    """
    获取成绩列表
    查询参数:
    - student_id: 学生ID（可选）
    - subject: 科目（可选）
    - page: 页码（默认1）
    - per_page: 每页数量（默认20）
    """
    try:
        teacher_id = payload.get('user_id')
        student_id = request.args.get('student_id', type=int)
        subject = request.args.get('subject')
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 20, type=int)
        
        # 构建查询
        query = StudentScore.query.filter_by(teacher_id=teacher_id, deleted_at=None)
        
        if student_id:
            query = query.filter_by(student_id=student_id)
        
        if subject:
            query = query.filter_by(subject=subject)
        
        # 分页
        paginated = query.order_by(StudentScore.exam_date.desc()).paginate(
            page=page,
            per_page=per_page
        )
        
        scores = [score.to_dict() for score in paginated.items]
        
        return jsonify({
            'code': 200,
            'message': '获取成功',
            'data': {
                'scores': scores,
                'total': paginated.total,
                'pages': paginated.pages,
                'current_page': page
            }
        }), 200
        
    except Exception as e:
        return jsonify({
            'code': 500,
            'message': f'获取失败: {str(e)}'
        }), 500


@grades_bp.route('/<int:score_id>', methods=['GET'])
@token_required
def get_score(payload, score_id):
    """获取单条成绩信息"""
    try:
        teacher_id = payload.get('user_id')
        score = StudentScore.query.filter_by(
            id=score_id,
            teacher_id=teacher_id,
            deleted_at=None
        ).first()
        
        if not score:
            return jsonify({
                'code': 404,
                'message': '成绩不存在'
            }), 404
        
        return jsonify({
            'code': 200,
            'message': '获取成功',
            'data': score.to_dict()
        }), 200
        
    except Exception as e:
        return jsonify({
            'code': 500,
            'message': f'获取失败: {str(e)}'
        }), 500


@grades_bp.route('/<int:score_id>', methods=['PUT'])
@token_required
def update_score(payload, score_id):
    """
    更新成绩信息
    """
    try:
        teacher_id = payload.get('user_id')
        data = request.get_json()
        
        score = StudentScore.query.filter_by(
            id=score_id,
            teacher_id=teacher_id,
            deleted_at=None
        ).first()
        
        if not score:
            return jsonify({
                'code': 404,
                'message': '成绩不存在'
            }), 404
        
        # 更新可修改的字段
        if 'score' in data:
            score.score = float(data['score'])
        if 'full_score' in data:
            score.full_score = float(data['full_score'])
        if 'subject' in data:
            score.subject = data['subject']
        if 'exam_date' in data:
            score.exam_date = datetime.strptime(data['exam_date'], '%Y-%m-%d').date()
        if 'exam_number' in data:
            score.exam_number = int(data['exam_number'])
        if 'notes' in data:
            score.notes = data['notes']
        
        db.session.commit()
        
        return jsonify({
            'code': 200,
            'message': '更新成功',
            'data': score.to_dict()
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'code': 500,
            'message': f'更新失败: {str(e)}'
        }), 500


@grades_bp.route('/<int:score_id>', methods=['DELETE'])
@token_required
def delete_score(payload, score_id):
    """删除成绩（软删除）"""
    try:
        teacher_id = payload.get('user_id')
        score = StudentScore.query.filter_by(
            id=score_id,
            teacher_id=teacher_id,
            deleted_at=None
        ).first()
        
        if not score:
            return jsonify({
                'code': 404,
                'message': '成绩不存在'
            }), 404
        
        score.deleted_at = datetime.utcnow()
        db.session.commit()
        
        return jsonify({
            'code': 200,
            'message': '删除成功'
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'code': 500,
            'message': f'删除失败: {str(e)}'
        }), 500


# ==================== 统计分析 ====================

@grades_bp.route('/statistics', methods=['GET'])
@token_required
def get_statistics(payload):
    """
    获取成绩统计数据
    查询参数:
    - subject: 科目（可选，如果不指定则统计所有科目）
    - student_id: 学生ID（可选）
    """
    try:
        teacher_id = payload.get('user_id')
        subject = request.args.get('subject')
        student_id = request.args.get('student_id', type=int)
        
        # 构建查询
        query = StudentScore.query.filter_by(teacher_id=teacher_id, deleted_at=None)
        
        if subject:
            query = query.filter_by(subject=subject)
        
        if student_id:
            query = query.filter_by(student_id=student_id)
        
        scores = query.all()
        
        if not scores:
            return jsonify({
                'code': 200,
                'message': '获取成功',
                'data': {
                    'count': 0,
                    'average': 0,
                    'max': 0,
                    'min': 0,
                    'std_dev': 0,
                    'distribution': {}
                }
            }), 200
        
        # 提取分数列表
        score_values = [s.score for s in scores]
        score_array = np.array(score_values)
        
        # 计算统计数据
        stats = {
            'count': len(scores),
            'average': round(float(np.mean(score_array)), 2),
            'max': float(np.max(score_array)),
            'min': float(np.min(score_array)),
            'std_dev': round(float(np.std(score_array)), 2),
            'median': round(float(np.median(score_array)), 2),
            'q1': round(float(np.percentile(score_array, 25)), 2),
            'q3': round(float(np.percentile(score_array, 75)), 2),
        }
        
        # 计算分数段分布
        distribution = {
            'A (90-100)': len([s for s in score_values if s >= 90]),
            'B (80-89)': len([s for s in score_values if 80 <= s < 90]),
            'C (70-79)': len([s for s in score_values if 70 <= s < 80]),
            'D (60-69)': len([s for s in score_values if 60 <= s < 70]),
            'F (<60)': len([s for s in score_values if s < 60]),
        }
        
        stats['distribution'] = distribution
        
        return jsonify({
            'code': 200,
            'message': '获取成功',
            'data': stats
        }), 200
        
    except Exception as e:
        return jsonify({
            'code': 500,
            'message': f'获取统计数据失败: {str(e)}'
        }), 500


@grades_bp.route('/trends', methods=['GET'])
@token_required
def get_trends(payload):
    """
    获取成绩变化趋势
    查询参数:
    - student_id: 学生ID（可选）
    - subject: 科目（可选）
    """
    try:
        teacher_id = payload.get('user_id')
        student_id = request.args.get('student_id', type=int)
        subject = request.args.get('subject')
        
        # 构建查询
        query = StudentScore.query.filter_by(teacher_id=teacher_id, deleted_at=None)
        
        if student_id:
            query = query.filter_by(student_id=student_id)
        
        if subject:
            query = query.filter_by(subject=subject)
        
        # 按考试日期排序
        scores = query.order_by(StudentScore.exam_date.asc()).all()
        
        if not scores:
            return jsonify({
                'code': 200,
                'message': '获取成功',
                'data': {
                    'dates': [],
                    'scores': [],
                    'average_line': []
                }
            }), 200
        
        # 按学生和科目分组
        trends_data = {}
        for score in scores:
            key = f"{score.student_id}_{score.subject}"
            if key not in trends_data:
                trends_data[key] = {
                    'student_name': User.query.get(score.student_id).fullname,
                    'subject': score.subject,
                    'dates': [],
                    'scores': []
                }
            trends_data[key]['dates'].append(score.exam_date.isoformat())
            trends_data[key]['scores'].append(score.score)
        
        # 转换为列表
        trends = list(trends_data.values())
        
        return jsonify({
            'code': 200,
            'message': '获取成功',
            'data': {
                'trends': trends,
                'count': len(trends)
            }
        }), 200
        
    except Exception as e:
        return jsonify({
            'code': 500,
            'message': f'获取趋势数据失败: {str(e)}'
        }), 500


@grades_bp.route('/student/<int:student_id>/summary', methods=['GET'])
@token_required
def get_student_summary(payload, student_id):
    """
    获取学生成绩汇总
    """
    try:
        teacher_id = payload.get('user_id')
        
        # 验证学生是否存在且是教师的学生
        student = User.query.get(student_id)
        if not student or student.identity != 'student':
            return jsonify({
                'code': 404,
                'message': '学生不存在'
            }), 404
        
        teacher_student = TeacherStudent.query.filter_by(
            teacher_id=teacher_id,
            student_id=student_id,
            deleted_at=None
        ).first()
        
        if not teacher_student:
            return jsonify({
                'code': 403,
                'message': '您没有权限查看该学生的成绩'
            }), 403
        
        # 获取该学生的所有成绩
        scores = StudentScore.query.filter_by(
            student_id=student_id,
            teacher_id=teacher_id,
            deleted_at=None
        ).all()
        
        # 按科目统计
        subject_stats = {}
        for score in scores:
            if score.subject not in subject_stats:
                subject_stats[score.subject] = {
                    'subject': score.subject,
                    'scores': [],
                    'count': 0,
                    'average': 0,
                    'max': 0,
                    'min': 0
                }
            
            subject_stats[score.subject]['scores'].append(score.score)
            subject_stats[score.subject]['count'] += 1
        
        # 计算统计值
        for subject, data in subject_stats.items():
            scores_array = np.array(data['scores'])
            data['average'] = round(float(np.mean(scores_array)), 2)
            data['max'] = float(np.max(scores_array))
            data['min'] = float(np.min(scores_array))
            data['scores'] = [float(s) for s in data['scores']]
        
        # 总体统计
        all_scores = np.array([s.score for s in scores])
        overall_stats = {
            'total_count': len(scores),
            'average': round(float(np.mean(all_scores)), 2) if len(scores) > 0 else 0,
            'max': float(np.max(all_scores)) if len(scores) > 0 else 0,
            'min': float(np.min(all_scores)) if len(scores) > 0 else 0,
        }
        
        return jsonify({
            'code': 200,
            'message': '获取成功',
            'data': {
                'student': {
                    'id': student.id,
                    'name': student.fullname,
                    'username': student.username
                },
                'overall': overall_stats,
                'subjects': list(subject_stats.values())
            }
        }), 200
        
    except Exception as e:
        return jsonify({
            'code': 500,
            'message': f'获取汇总失败: {str(e)}'
        }), 500


@grades_bp.route('/subjects', methods=['GET'])
@token_required
def get_subjects(payload):
    """获取教师教授的所有科目"""
    try:
        teacher_id = payload.get('user_id')
        
        # 获取该教师输入过的所有科目
        subjects = db.session.query(StudentScore.subject).distinct().filter_by(
            teacher_id=teacher_id,
            deleted_at=None
        ).all()
        
        subject_list = [s[0] for s in subjects]
        
        return jsonify({
            'code': 200,
            'message': '获取成功',
            'data': subject_list
        }), 200
        
    except Exception as e:
        return jsonify({
            'code': 500,
            'message': f'获取科目列表失败: {str(e)}'
        }), 500


@grades_bp.route('/students', methods=['GET'])
@token_required
def get_students(payload):
    """获取教师教授的所有学生"""
    try:
        teacher_id = payload.get('user_id')
        
        # 获取该教师的所有学生
        teacher_students = TeacherStudent.query.filter_by(
            teacher_id=teacher_id,
            deleted_at=None
        ).all()
        
        students = []
        for ts in teacher_students:
            student = User.query.get(ts.student_id)
            if student:
                students.append({
                    'id': student.id,
                    'name': student.fullname,
                    'username': student.username
                })
        
        return jsonify({
            'code': 200,
            'message': '获取成功',
            'data': students
        }), 200
        
    except Exception as e:
        return jsonify({
            'code': 500,
            'message': f'获取学生列表失败: {str(e)}'
        }), 500
